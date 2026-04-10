import io
import streamlit as st
import pandas as pd
import pulp
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Workforce Optimizer", layout="wide")

st.markdown(
    "<style>[data-testid='stSidebar']{min-width:320px;max-width:420px}</style>",
    unsafe_allow_html=True,
)

REQUIRED_COLS = {
    "Employees": {"employee", "capacity", "skills", "hourly_rate ($)"},
    "Projects":  {"project", "reimbursable", "max_total", "budget ($)"},
    "Tasks":     {"project", "task_id", "task_type", "min_hours"},
}


def build_blank_template() -> bytes:
    def write_sheet(ws, headers, rows, notes, widths):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(r, c, v)
        for c, n in enumerate(notes, 1):
            cell = ws.cell(len(rows) + 2, c, n)
            cell.font = Font(italic=True, size=9)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    write_sheet(ws,
        ["employee", "capacity", "skills", "hourly_rate ($)", "employee_type"],
        [["E001", 40, "A,B", 67.50, "Senior"], ["E002", 30, "B,C,D", 55.00, "Mid-Level"]],
        ["Unique ID", "Max hrs/week", "Comma-sep A-E", "Hourly rate ($)", "Junior/Mid-Level/Senior"],
        [16, 14, 18, 18, 16],
    )

    ws = wb.create_sheet("Projects")
    write_sheet(ws,
        ["project", "reimbursable", "max_total", "budget ($)"],
        [["P001", True, 90, 12000], ["P002", False, None, 8500]],
        ["Unique ID", "TRUE=billable FALSE=internal", "Max hours (blank if non-reimb.)", "Dollar budget"],
        [16, 16, 14, 14],
    )

    ws = wb.create_sheet("Tasks")
    write_sheet(ws,
        ["project", "task_id", "task_type", "min_hours", "Urgency"],
        [["P001", "T001", "A", 10, None], ["P001", "T002", "B", 14, None]],
        ["Match a project ID", "Unique task ID", "Skill A/B/C/D/E", "Min hours", "Optional"],
        [14, 14, 14, 14, 12],
    )

    ws = wb.create_sheet("Seeds")
    write_sheet(ws,
        ["employee", "task_id", "seed_hours"],
        [["E001", "T001", 8], ["E002", "T002", 12]],
        ["Must match Employees sheet", "Must match Tasks sheet", "Expected hours (best guess)"],
        [18, 16, 16],
    )

    ws = wb.create_sheet("Data Dictionary")
    dict_headers = ["Sheet", "Column", "Type", "Description"]
    dict_rows = [
        ("Employees", "employee",        "ID",       "Unique employee identifier"),
        ("Employees", "capacity",        "Integer",  "Max weekly hours available"),
        ("Employees", "skills",          "String",   "Comma-separated skill types A-E"),
        ("Employees", "hourly_rate ($)", "Float",    "Hourly billing/wage rate"),
        ("Employees", "employee_type",   "Category", "Junior / Mid-Level / Senior"),
        ("Projects",  "project",         "ID",       "Unique project identifier"),
        ("Projects",  "reimbursable",    "Boolean",  "TRUE = client-billable"),
        ("Projects",  "max_total",       "Integer",  "Max hours cap (blank if non-reimbursable)"),
        ("Projects",  "budget ($)",      "Float",    "Dollar budget ceiling for wage costs"),
        ("Tasks",     "project",         "ID",       "Must match a project ID"),
        ("Tasks",     "task_id",         "ID",       "Unique task identifier"),
        ("Tasks",     "task_type",       "Category", "Required skill type (A-E)"),
        ("Tasks",     "min_hours",       "Integer",  "Minimum hours to complete task"),
        ("Tasks",     "Urgency",         "Optional", "Not used by optimizer"),
        ("Seeds",     "employee",        "ID",       "Must match an employee ID — sheet is optional"),
        ("Seeds",     "task_id",         "ID",       "Must match a task ID"),
        ("Seeds",     "seed_hours",      "Float",    "Expected hours for this employee on this task"),
    ]
    for c, h in enumerate(dict_headers, 1):
        ws.cell(1, c, h).font = Font(bold=True)
    for r, row in enumerate(dict_rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    for i, w in enumerate([14, 18, 12, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def validate_template(xl):
    missing_sheets = {"Employees", "Projects", "Tasks"} - set(xl.sheet_names)
    if missing_sheets:
        return False, f"Missing sheet(s): {', '.join(missing_sheets)}"
    errors = []
    for sheet, required in REQUIRED_COLS.items():
        cols = set(pd.read_excel(xl, sheet_name=sheet, nrows=0).columns.str.strip())
        missing = required - cols
        if missing:
            errors.append(f"{sheet} missing: {', '.join(missing)}")
    return (False, "\n".join(errors)) if errors else (True, None)


def load_data(xl):
    def strip_cols(df):
        df.columns = df.columns.str.strip()
        return df

    emp_df = strip_cols(pd.read_excel(xl, sheet_name="Employees"))
    employees = [
        {
            "id":       str(r["employee"]).strip(),
            "capacity": int(r["capacity"]),
            "skills":   [s.strip() for s in str(r["skills"]).split(",") if s.strip()],
            "rate":     float(r["hourly_rate ($)"]),
            "type":     str(r.get("employee_type", "")).strip(),
        }
        for _, r in emp_df.iterrows()
    ]

    proj_df = strip_cols(pd.read_excel(xl, sheet_name="Projects"))
    projects = [
        {
            "id":          str(r["project"]).strip(),
            "reimbursable": bool(r["reimbursable"]),
            "maxTotal":    None if pd.isna(r["max_total"]) else int(r["max_total"]),
            "budget":      None if pd.isna(r["budget ($)"]) else float(r["budget ($)"]),
        }
        for _, r in proj_df.iterrows()
    ]

    task_df = strip_cols(pd.read_excel(xl, sheet_name="Tasks"))
    tasks = [
        {
            "project":  str(r["project"]).strip(),
            "id":       str(r["task_id"]).strip(),
            "type":     str(r["task_type"]).strip(),
            "minHours": int(r["min_hours"]),
        }
        for _, r in task_df.iterrows()
    ]

    seeds = {}
    if "Seeds" in xl.sheet_names:
        seed_df = strip_cols(pd.read_excel(xl, sheet_name="Seeds"))
        for _, r in seed_df.iterrows():
            emp_id  = str(r["employee"]).strip()
            task_id = str(r["task_id"]).strip()
            hours   = float(r["seed_hours"])
            if hours > 0:
                seeds[(emp_id, task_id)] = hours

    return employees, projects, tasks, seeds


def run_optimizer(employees, tasks, projects, seeds=None, seed_weight=1.0):
    seeds = seeds or {}
    emp_map  = {e["id"]: e for e in employees}
    task_map = {t["id"]: t for t in tasks}
    proj_map = {p["id"]: p for p in projects}

    pairs = [(e["id"], t["id"]) for e in employees for t in tasks
             if t["type"] in e["skills"]]

    prob = pulp.LpProblem("wf", pulp.LpMinimize)
    x = {(i, j): pulp.LpVariable(f"x_{i}_{j}", lowBound=0) for i, j in pairs}
    s = {t["id"]: pulp.LpVariable(f"s_{t['id']}", lowBound=0) for t in tasks}

    # Seed deviation variables
    seed_pairs = [(i, j) for (i, j) in seeds if (i, j) in x]
    dp = {k: pulp.LpVariable(f"dp_{k[0]}_{k[1]}", lowBound=0) for k in seed_pairs}
    dm = {k: pulp.LpVariable(f"dm_{k[0]}_{k[1]}", lowBound=0) for k in seed_pairs}
    for k in seed_pairs:
        prob += x[k] - dp[k] + dm[k] == seeds[k]

    prob += (
        pulp.lpSum(
            (2 if proj_map.get(task_map[j]["project"], {}).get("reimbursable") else 1) * s[j]
            for j in s
        )
        + seed_weight * pulp.lpSum(dp[k] + dm[k] for k in seed_pairs)
    )

    for e in employees:
        ep = [j for i, j in pairs if i == e["id"]]
        if ep:
            prob += pulp.lpSum(x[(e["id"], j)] for j in ep) <= e["capacity"]

    for t in tasks:
        tp = [x[(i, t["id"])] for i, j in pairs if j == t["id"]]
        if tp:
            prob += pulp.lpSum(tp) + s[t["id"]] >= t["minHours"]
        else:
            prob += s[t["id"]] >= t["minHours"]

    for p in projects:
        if p["reimbursable"] and p["maxTotal"]:
            pp = [(i, j) for i, j in pairs if task_map[j]["project"] == p["id"]]
            if pp:
                prob += pulp.lpSum(x[k] for k in pp) <= p["maxTotal"]

    for p in projects:
        if p["budget"]:
            pp = [(i, j) for i, j in pairs if task_map[j]["project"] == p["id"]]
            if pp:
                prob += pulp.lpSum(x[(i, j)] * emp_map[i]["rate"] for i, j in pp) <= p["budget"]

    prob.solve(pulp.PULP_CBC_CMD(msg=0))

    load    = {e["id"]: 0.0 for e in employees}
    p_hours = {p["id"]: 0.0 for p in projects}
    p_cost  = {p["id"]: 0.0 for p in projects}
    ta      = {t["id"]: {} for t in tasks}

    for i, j in pairs:
        v = pulp.value(x[(i, j)])
        if v and v > 0.01:
            ta[j][i] = v
            load[i]                          += v
            p_hours[task_map[j]["project"]]  += v
            p_cost[task_map[j]["project"]]   += v * emp_map[i]["rate"]

    asgn = {}
    for t in tasks:
        assigned = ta[t["id"]]
        if not assigned:
            asgn[t["id"]] = {"employee": None, "hours": 0}
        else:
            primary = max(assigned, key=assigned.get)
            hours   = sum(assigned.values())
            asgn[t["id"]] = {
                "employee": primary,
                "hours":    round(hours, 1),
                "partial":  hours < t["minHours"] - 0.5,
                "seeded":   (primary, t["id"]) in seed_pairs,
            }

    return {
        "asgn":    asgn,
        "load":    {k: round(v, 1) for k, v in load.items()},
        "p_hours": {k: round(v, 1) for k, v in p_hours.items()},
        "p_cost":  {k: round(v, 2) for k, v in p_cost.items()},
        "status":  pulp.LpStatus[prob.status],
    }


# ── Page header ──────────────────────────────────────────────────────────────
st.markdown("## Workforce Optimizer")
st.caption("Upload your data template in the sidebar to run the optimizer.")

with st.sidebar:
    st.header("Data Import")
    st.markdown("Upload an Excel file with three sheets: **Employees**, **Projects**, and **Tasks**.")

    try:
        st.download_button(
            "Download blank template",
            data=build_blank_template(),
            file_name="workforce_optimizer_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception:
        st.caption("Template download unavailable.")

    try:
        with open("Capstone Budget Data.xlsx", "rb") as f:
            st.download_button(
                "Download example dataset",
                data=f.read(),
                file_name="Capstone Budget Data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    except Exception:
        st.caption("Example dataset unavailable.")

    st.divider()
    uploaded_file = st.file_uploader("Choose Excel file (.xlsx)", type=["xlsx"])
    st.divider()
    seed_weight = st.slider(
        "Seed adherence",
        min_value=0.0, max_value=5.0, value=1.0, step=0.5,
        help="0 = ignore seeds, 5 = follow seeds very closely",
    )

if uploaded_file is None:
    st.info("Upload your Excel template in the sidebar to get started.")
    st.stop()

try:
    xl = pd.ExcelFile(uploaded_file)
except Exception as e:
    st.error(f"Could not read file: {e}")
    st.stop()

valid, err = validate_template(xl)
if not valid:
    st.error(f"Invalid template:\n\n{err}")
    st.stop()

try:
    EMPLOYEES, PROJECTS, TASKS, SEEDS = load_data(xl)
except Exception as e:
    st.error(f"Error loading data: {e}")
    st.stop()

seed_info = f" · {len(SEEDS)} seeds" if SEEDS else ""
st.sidebar.success(f"{len(EMPLOYEES)} employees · {len(PROJECTS)} projects · {len(TASKS)} tasks{seed_info}")

cache_key = f"{uploaded_file.name}_{uploaded_file.size}_{seed_weight}"
if st.session_state.get("_cache_key") != cache_key:
    with st.spinner("Running optimizer..."):
        st.session_state["opt_result"]  = run_optimizer(EMPLOYEES, TASKS, PROJECTS, SEEDS, seed_weight)
        st.session_state["_cache_key"] = cache_key

opt     = st.session_state["opt_result"]
asgn    = opt["asgn"]
load    = opt["load"]
p_hours = opt["p_hours"]
p_cost  = opt["p_cost"]

proj_map = {p["id"]: p for p in PROJECTS}
emp_map  = {e["id"]: e for e in EMPLOYEES}

total_cap  = sum(e["capacity"] for e in EMPLOYEES)
total_load = round(sum(load.values()), 1)
util       = round(total_load / total_cap * 100) if total_cap else 0
n_ok   = sum(1 for a in asgn.values() if a.get("employee") and not a.get("partial"))
n_fail = sum(1 for a in asgn.values() if not a.get("employee"))
n_part = sum(1 for a in asgn.values() if a.get("partial"))

tasks_by_emp = {e["id"]: [] for e in EMPLOYEES}
for t in TASKS:
    a = asgn.get(t["id"], {})
    if a.get("employee"):
        tasks_by_emp[a["employee"]].append({**t, **a})

# ── KPIs ─────────────────────────────────────────────────────────────────────
c1, c2, c3 = st.columns(3)
c1.metric("Utilization",   f"{util}%")
c2.metric("Tasks Covered", f"{n_ok} / {len(TASKS)}")
c3.metric("Unfilled",      n_fail + n_part)

if opt["status"] not in ("Optimal", "Not Solved"):
    st.warning(f"Solver status: {opt['status']} — results may be incomplete.")

st.divider()

tabs = st.tabs(["Dashboard", "Employees", "Projects", "Assignments", "Skills"])


# ── Dashboard tab ─────────────────────────────────────────────────────────────
with tabs[0]:
    reimb = sum(1 for p in PROJECTS if p["reimbursable"])
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Employees",   len(EMPLOYEES), f"{total_cap} h/week capacity")
    c2.metric("Projects",    len(PROJECTS),  f"{reimb} reimbursable")
    c3.metric("Tasks",       f"{n_ok} / {len(TASKS)}", f"{n_fail + n_part} need attention" if n_fail + n_part else "All covered")
    c4.metric("Utilization", f"{util}%",     f"{total_load} of {total_cap} h used")

    st.subheader("Employee Utilization")
    st.dataframe(
        pd.DataFrame([{
            "Employee":      e["id"],
            "Type":          e["type"],
            "Rate ($/h)":    f"${e['rate']:.2f}",
            "Skills":        ", ".join(e["skills"]),
            "Assigned (h)":  load.get(e["id"], 0),
            "Capacity (h)":  e["capacity"],
            "Utilization %": round(load.get(e["id"], 0) / e["capacity"] * 100) if e["capacity"] else 0,
        } for e in EMPLOYEES]),
        use_container_width=True, hide_index=True,
        column_config={"Utilization %": st.column_config.ProgressColumn(
            "Utilization %", min_value=0, max_value=100, format="%d%%"
        )},
    )

    st.subheader("Skill Demand vs Supply")
    skill_rows = []
    all_skill_types = sorted(set(t["type"] for t in TASKS))
    skill_demand = {sk: sum(t["minHours"] for t in TASKS if t["type"] == sk)
                    for sk in all_skill_types}

    for sk in all_skill_types:
        dem = skill_demand[sk]
        sup = 0.0
        for e in EMPLOYEES:
            if sk not in e["skills"]:
                continue
            rel_skills = [s for s in e["skills"] if s in skill_demand]
            total_rel_demand = sum(skill_demand[s] for s in rel_skills)
            if total_rel_demand > 0:
                sup += e["capacity"] * (dem / total_rel_demand)
            else:
                sup += e["capacity"] / max(len(e["skills"]), 1)
        ratio = dem / sup if sup else 99
        skill_rows.append({
            "Skill":      sk,
            "Employees":  sum(1 for e in EMPLOYEES if sk in e["skills"]),
            "Tasks":      sum(1 for t in TASKS if t["type"] == sk),
            "Demand (h)": dem,
            "Supply (h)": sup,
            "Coverage %": round(min(sup / dem, 1) * 100) if dem else 100,
            "Status":     "CRITICAL" if ratio > 1.3 else "TIGHT" if ratio > 1.0
                          else "HEALTHY" if ratio > 0.6 else "SURPLUS",
        })
    st.dataframe(pd.DataFrame(skill_rows), use_container_width=True, hide_index=True,
        column_config={"Coverage %": st.column_config.ProgressColumn(
            "Coverage %", min_value=0, max_value=100, format="%d%%"
        )},
    )

    st.subheader("Project Summary")
    proj_sum_rows = []
    for p in PROJECTS:
        dem = sum(t["minHours"] for t in TASKS if t["project"] == p["id"])
        asn = p_hours.get(p["id"], 0)
        cost = p_cost.get(p["id"], 0)
        over = (p["maxTotal"] and asn > p["maxTotal"]) or (p["budget"] and cost > p["budget"])
        proj_sum_rows.append({
            "Project":       p["id"],
            "Type":          "Reimbursable" if p["reimbursable"] else "Non-Reimb.",
            "Demand (h)":    dem,
            "Assigned (h)":  round(asn, 1),
            "Hour Cap":      f"{p['maxTotal']} h" if p["maxTotal"] else "—",
            "Wage Cost ($)": f"${cost:,.0f}",
            "Budget ($)":    f"${p['budget']:,.0f}" if p["budget"] else "—",
            "Coverage %":    round(asn / dem * 100) if dem else 100,
            "Status":        "Over Budget" if over else "Fulfilled" if asn >= dem else "Partial",
        })
    st.dataframe(pd.DataFrame(proj_sum_rows), use_container_width=True, hide_index=True,
        column_config={"Coverage %": st.column_config.ProgressColumn(
            "Coverage %", min_value=0, max_value=100, format="%d%%"
        )},
    )


# ── Employees tab ─────────────────────────────────────────────────────────────
with tabs[1]:
    st.subheader("Employee Roster")
    for e in EMPLOYEES:
        hrs  = load.get(e["id"], 0)
        util_e = round(hrs / e["capacity"] * 100) if e["capacity"] else 0
        label = (f"**{e['id']}** [{e['type']}]  |  Skills: {', '.join(e['skills'])}"
                 f"  |  ${e['rate']:.2f}/h  |  {util_e}% utilized ({hrs}/{e['capacity']} h)")
        with st.expander(label):
            et = tasks_by_emp.get(e["id"], [])
            if not et:
                st.caption("No tasks assigned.")
            else:
                st.dataframe(pd.DataFrame([{
                    "Task":         t["id"],
                    "Project":      t["project"],
                    "Skill":        t["type"],
                    "Min Hrs":      t["minHours"],
                    "Assigned Hrs": round(t.get("hours", t["minHours"]), 1),
                    "Status":       "Partial" if t.get("partial") else "OK",
                } for t in et]), use_container_width=True, hide_index=True)


# ── Projects tab ──────────────────────────────────────────────────────────────
with tabs[2]:
    st.subheader("Project Overview")
    for p in PROJECTS:
        pt   = [t for t in TASKS if t["project"] == p["id"]]
        dem  = sum(t["minHours"] for t in pt)
        asn  = p_hours.get(p["id"], 0)
        cost = p_cost.get(p["id"], 0)
        over = (p["maxTotal"] and asn > p["maxTotal"]) or (p["budget"] and cost > p["budget"])
        status_label = "Over Budget" if over else ("Fulfilled" if asn >= dem else "Partial")
        header = (f"**{p['id']}**  |  {'Reimbursable' if p['reimbursable'] else 'Non-Reimbursable'}"
                  f"  |  {status_label}  |  {round(asn, 1)}/{dem} h  |  ${cost:,.0f}")

        with st.expander(header):
            if p["maxTotal"]:
                st.progress(min(asn / p["maxTotal"], 1.0),
                            text=f"Hour budget: {round(asn,1)} / {p['maxTotal']} h")
            if p["budget"]:
                st.progress(min(cost / p["budget"], 1.0),
                            text=f"Wage budget: ${cost:,.0f} / ${p['budget']:,.0f}")

            st.markdown("##### Cost Breakdown")
            cb = {}
            for t in pt:
                a = asgn.get(t["id"], {})
                eid = a.get("employee")
                if eid:
                    cb.setdefault(eid, {"tasks": 0, "hours": 0.0})
                    cb[eid]["tasks"]  += 1
                    cb[eid]["hours"]  += a.get("hours", 0)

            if cb:
                cb_rows = [{
                    "Employee":   eid,
                    "Tasks":      v["tasks"],
                    "Hours":      round(v["hours"], 1),
                    "Rate ($/h)": f"${emp_map[eid]['rate']:.2f}",
                    "Cost ($)":   f"${v['hours'] * emp_map[eid]['rate']:,.2f}",
                } for eid, v in cb.items()]
                cb_rows.append({
                    "Employee": "TOTAL",
                    "Tasks":    sum(r["Tasks"] for r in cb_rows),
                    "Hours":    round(sum(r["Hours"] for r in cb_rows), 1),
                    "Rate ($/h)": "—",
                    "Cost ($)": f"${cost:,.2f}",
                })
                st.dataframe(pd.DataFrame(cb_rows), use_container_width=True, hide_index=True)
            else:
                st.caption("No tasks assigned to this project.")

            st.markdown("##### Task Details")
            st.dataframe(pd.DataFrame([{
                "Task":         t["id"],
                "Skill":        t["type"],
                "Min Hrs":      t["minHours"],
                "Assigned Hrs": round(asgn.get(t["id"], {}).get("hours", 0), 1),
                "Assigned To":  asgn.get(t["id"], {}).get("employee") or "—",
                "Status":       "Partial"    if asgn.get(t["id"], {}).get("partial")
                                else "OK"    if asgn.get(t["id"], {}).get("employee")
                                else "Unassigned",
            } for t in pt]), use_container_width=True, hide_index=True)


# ── Assignments tab ───────────────────────────────────────────────────────────
with tabs[3]:
    st.subheader("Full Task Assignment Matrix")
    rows = []
    for t in TASKS:
        a   = asgn.get(t["id"], {})
        eid = a.get("employee")
        hrs = a.get("hours", 0)
        cost_val = f"${hrs * emp_map[eid]['rate']:,.2f}" if eid and eid in emp_map else "—"
        rows.append({
            "Task":         t["id"],
            "Project":      f"{t['project']} ({'R' if proj_map.get(t['project'], {}).get('reimbursable') else 'N'})",
            "Skill":        t["type"],
            "Min Hours":    t["minHours"],
            "Assigned Hrs": round(hrs, 1),
            "Assigned To":  eid or "—",
            "Cost ($)":     cost_val,
            "Status":       "Failed"    if not eid
                            else "Partial" if a.get("partial")
                            else "OK",
            "Seeded":      "Yes" if a.get("seeded") else "—",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ── Skills tab ────────────────────────────────────────────────────────────────
with tabs[4]:
    st.subheader("Skills Overview")

    all_skills_in_tasks = sorted(set(t["type"] for t in TASKS))
    all_skills_in_emps  = sorted(set(sk for e in EMPLOYEES for sk in e["skills"]))
    all_skills          = sorted(set(all_skills_in_tasks) | set(all_skills_in_emps))

    skill_demand_hrs = {sk: sum(t["minHours"] for t in TASKS if t["type"] == sk) for sk in all_skills}
    skill_holders    = {sk: [e["id"] for e in EMPLOYEES if sk in e["skills"]] for sk in all_skills}
    skill_task_count = {sk: sum(1 for t in TASKS if t["type"] == sk) for sk in all_skills}

    critical_gaps  = [sk for sk in all_skills_in_tasks if not skill_holders.get(sk)]
    single_holder  = [sk for sk in all_skills_in_tasks if len(skill_holders.get(sk, [])) == 1]
    unused_skills  = [sk for sk in all_skills_in_emps if sk not in all_skills_in_tasks]

    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Critical Gaps",     len(critical_gaps),
                 f"No coverage: {', '.join(critical_gaps)}" if critical_gaps else "None")
    col_b.metric("Single-Owner Skills", len(single_holder),
                 f"Only 1 holder: {', '.join(single_holder)}" if single_holder else "None")
    col_c.metric("Unused Skills",     len(unused_skills),
                 f"Not needed: {', '.join(unused_skills)}" if unused_skills else "None")

    st.divider()

    # Employee skill breakdown
    st.subheader("Employee Skills")
    emp_skill_rows = []
    for e in EMPLOYEES:
        util_pct      = round(load.get(e["id"], 0) / e["capacity"] * 100) if e["capacity"] else 0
        active        = [sk for sk in e["skills"] if sk in all_skills_in_tasks]
        inactive      = [sk for sk in e["skills"] if sk not in all_skills_in_tasks]
        missing       = [sk for sk in all_skills_in_tasks if sk not in e["skills"]]
        train_priority = (
            "High"   if any(sk in critical_gaps or sk in single_holder for sk in missing)
            else "Medium" if missing
            else "Low"
        )
        emp_skill_rows.append({
            "Employee":          e["id"],
            "Level":             e["type"],
            "Utilization %":     util_pct,
            "Current Skills":    ", ".join(active) if active else "—",
            "Unused Skills":     ", ".join(inactive) if inactive else "—",
            "Missing Skills":    ", ".join(missing) if missing else "None",
            "Training Priority": train_priority,
        })

    st.dataframe(
        pd.DataFrame(emp_skill_rows),
        use_container_width=True, hide_index=True,
        column_config={"Utilization %": st.column_config.ProgressColumn(
            "Utilization %", min_value=0, max_value=100, format="%d%%"
        )},
    )

    st.divider()

    # Employee x Skill matrix
    st.subheader("Employee x Skill Matrix")
    st.caption("Yes = has skill   |   No = missing   |   Unused = has skill but no task requires it")
    matrix_rows = []
    for e in EMPLOYEES:
        row = {"Employee": e["id"], "Level": e["type"]}
        for sk in all_skills:
            if sk in e["skills"]:
                row[f"Skill {sk}"] = "Yes" if sk in all_skills_in_tasks else "Unused"
            else:
                row[f"Skill {sk}"] = "No" if sk in all_skills_in_tasks else "—"
        matrix_rows.append(row)
    st.dataframe(pd.DataFrame(matrix_rows), use_container_width=True, hide_index=True)

    st.divider()

    # Training and recruiting side by side
    col_train, col_recruit = st.columns(2)

    with col_train:
        st.subheader("Training Candidates")
        st.caption("Under-utilized employees who could learn high-priority skills.")
        train_recs = []
        for sk in critical_gaps + single_holder:
            for e in EMPLOYEES:
                if sk in e["skills"]:
                    continue
                util_e = round(load.get(e["id"], 0) / e["capacity"] * 100) if e["capacity"] else 0
                if util_e < 80:
                    train_recs.append({
                        "Employee":       e["id"],
                        "Level":          e["type"],
                        "Skill to Learn": sk,
                        "Reason":         "No coverage" if sk in critical_gaps else "Only 1 holder",
                        "Utilization %":  util_e,
                        "Urgency":        "Urgent" if sk in critical_gaps else "Important",
                    })
        if train_recs:
            st.dataframe(pd.DataFrame(train_recs), use_container_width=True, hide_index=True,
                column_config={"Utilization %": st.column_config.ProgressColumn(
                    "Utilization %", min_value=0, max_value=100, format="%d%%"
                )})
        else:
            st.info("No immediate training needs based on current gaps.")

    with col_recruit:
        st.subheader("Recruiting Needs")
        st.caption("Skills where staff capacity falls short of task demand.")
        recruit_recs = []
        for sk in all_skills_in_tasks:
            holders      = skill_holders.get(sk, [])
            dem          = skill_demand_hrs.get(sk, 0)
            total_cap_sk = sum(e["capacity"] for e in EMPLOYEES if sk in e["skills"])
            if not holders:
                recruit_recs.append({
                    "Skill":            sk,
                    "Current Holders":  0,
                    "Demand (h)":       dem,
                    "Capacity Gap (h)": dem,
                    "Priority":         "Immediate",
                })
            elif total_cap_sk < dem:
                recruit_recs.append({
                    "Skill":            sk,
                    "Current Holders":  len(holders),
                    "Demand (h)":       dem,
                    "Capacity Gap (h)": round(dem - total_cap_sk, 1),
                    "Priority":         "High" if (dem - total_cap_sk) > 20 else "Medium",
                })
        if recruit_recs:
            st.dataframe(pd.DataFrame(recruit_recs), use_container_width=True, hide_index=True)
        else:
            st.info("Current staff capacity meets all skill demands.")

    st.divider()

    # Risk analysis
    st.subheader("Coverage Risk")
    st.caption("Skills with few holders are a business risk if someone becomes unavailable.")
    risk_rows = []
    for sk in all_skills_in_tasks:
        holders = skill_holders.get(sk, [])
        n       = len(holders)
        risk_rows.append({
            "Skill":           sk,
            "Holders":         n,
            "Holder IDs":      ", ".join(holders) if holders else "—",
            "Task Demand (h)": skill_demand_hrs.get(sk, 0),
            "Risk Level":      "Critical" if n == 0 else "High" if n == 1 else "Medium" if n == 2 else "Low",
            "Notes":           "No coverage" if n == 0
                               else f"All demand on {holders[0]} alone" if n == 1
                               else "Losing one person creates strain" if n == 2
                               else f"{n} employees can cover this",
        })
    st.dataframe(pd.DataFrame(risk_rows), use_container_width=True, hide_index=True)
