import io
import hashlib
import math
import streamlit as st
import pandas as pd
import pulp
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Workforce Optimizer", layout="wide")

# Widen the sidebar slightly beyond Streamlit's default max
st.markdown(
    "<style>[data-testid='stSidebar']{min-width:320px;max-width:420px}</style>",
    unsafe_allow_html=True,
)


def status_tier(status: str) -> str:
    if status in {"Fulfilled", "OK"}:
        return "Green"
    if status in {"Partial"}:
        return "Yellow"
    return "Red"


def format_supply_vs_demand_pct(sup: float, dem: float) -> str:
    """+76% = surplus vs task hours; -24% = shortfall."""
    if dem <= 0:
        return "—" if sup <= 0 else "No task demand"
    pct = round((sup / dem - 1) * 100)
    return f"+{pct}%" if pct > 0 else f"{pct}%"


def roster_skill_supply_hours(
    employees,
    skill: str,
    task_skill_types: set[str],
) -> float:
    """Capacity split only across skills that appear on tasks; result floored to whole hours."""
    if skill not in task_skill_types:
        return 0.0
    total = 0.0
    for e in employees:
        skills = e.get("skills") or []
        if skill not in skills:
            continue
        active = [s for s in skills if s in task_skill_types]
        if not active:
            continue
        total += e["capacity"] / len(active)
    return math.floor(total)


def status_style(v: str) -> str:
    if v == "Green":
        return "background-color: #d1fae5; color: #065f46; font-weight: 600;"
    if v == "Yellow":
        return "background-color: #fef3c7; color: #92400e; font-weight: 600;"
    if v == "Red":
        return "background-color: #fee2e2; color: #991b1b; font-weight: 600;"
    return ""

# Required columns for each sheet - used by validate_template
REQUIRED_COLS = {
    "Employees": {"employee", "capacity", "skills", "hourly_rate ($)"},
    "Projects":  {"project", "reimbursable", "max_total", "budget ($)"},
    "Tasks":     {"project", "task_id", "task_type", "min_hours"},
}


# Build and return an in-memory blank Excel template (.xlsx bytes).
# Each sheet gets a bold header row, two example rows, and a notes row. No color styling.
def build_blank_template() -> bytes:
    def write_sheet(ws, headers, rows, notes, widths):
        # Header row - bold
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        # Example data rows
        for r, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(r, c, v)
        # Notes row - italic, small
        for c, n in enumerate(notes, 1):
            cell = ws.cell(len(rows) + 2, c, n)
            cell.font = Font(italic=True, size=9)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    write_sheet(ws,
        ["employee", "capacity", "skills", "hourly_rate ($)", "employee_type"],
        [["E001", 40, "A,B", 67.50, "Senior"], ["E002", 30, "B,C,D", 55.00, "Mid-Level"]],
        ["Unique ID", "Max hrs/week", "Comma-sep A-E", "Hourly rate ($)", "Junior/Mid-Level/Senior"],
        [16, 14, 18, 18, 16],
    )

    ws = wb.create_sheet("Projects")
    write_sheet(ws,
        ["project", "reimbursable", "max_total", "budget ($)"],
        [["P001", True, 90, 12000], ["P002", False, None, 8500]],
        ["Unique ID", "TRUE=billable FALSE=internal", "Max hours (blank if non-reimb.)", "Dollar budget"],
        [16, 16, 14, 14],
    )

    ws = wb.create_sheet("Tasks")
    write_sheet(ws,
        ["project", "task_id", "task_type", "min_hours", "Urgency"],
        [["P001", "T001", "A", 10, None], ["P001", "T002", "B", 14, None]],
        ["Match a project ID", "Unique task ID", "Skill A/B/C/D/E", "Min hours", "Optional"],
        [14, 14, 14, 14, 12],
    )

    ws = wb.create_sheet("Data Dictionary")
    dict_headers = ["Sheet", "Column", "Type", "Description"]
    dict_rows = [
        ("Employees", "employee",        "ID",       "Unique employee identifier"),
        ("Employees", "capacity",        "Integer",  "Max weekly hours available"),
        ("Employees", "skills",          "String",   "Comma-separated skill types A-E"),
        ("Employees", "hourly_rate ($)", "Float",    "Hourly billing/wage rate"),
        ("Employees", "employee_type",   "Category", "Junior / Mid-Level / Senior"),
        ("Projects",  "project",         "ID",       "Unique project identifier"),
        ("Projects",  "reimbursable",    "Boolean",  "TRUE = client-billable"),
        ("Projects",  "max_total",       "Integer",  "Max hours cap (blank if non-reimbursable)"),
        ("Projects",  "budget ($)",      "Float",    "Dollar budget ceiling for wage costs"),
        ("Tasks",     "project",         "ID",       "Must match a project ID"),
        ("Tasks",     "task_id",         "ID",       "Unique task identifier"),
        ("Tasks",     "task_type",       "Category", "Required skill type (A-E)"),
        ("Tasks",     "min_hours",       "Integer",  "Minimum hours to complete task"),
        ("Tasks",     "Urgency",         "Optional", "Not used by optimizer"),
    ]
    for c, h in enumerate(dict_headers, 1):
        ws.cell(1, c, h).font = Font(bold=True)
    for r, row in enumerate(dict_rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    for i, w in enumerate([14, 18, 12, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# Check that the uploaded file has the required sheets and columns.
# Returns (True, None) on success or (False, error string) on failure.
def validate_template(xl):
    missing_sheets = {"Employees", "Projects", "Tasks"} - set(xl.sheet_names)
    if missing_sheets:
        return False, f"Missing sheet(s): {', '.join(missing_sheets)}"
    errors = []
    for sheet, required in REQUIRED_COLS.items():
        cols = set(pd.read_excel(xl, sheet_name=sheet, nrows=0).columns.str.strip())
        missing = required - cols
        if missing:
            errors.append(f"{sheet} missing: {', '.join(missing)}")
    return (False, "\n".join(errors)) if errors else (True, None)


# Parse the three sheets into lists of dicts for the optimizer and UI.
def load_data(xl):
    def to_bool(v):
        if pd.isna(v):
            return False
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in {"true", "t", "yes", "y", "1"}:
            return True
        if s in {"false", "f", "no", "n", "0", ""}:
            return False
        raise ValueError(f"Invalid reimbursable value: {v!r}. Use TRUE/FALSE, YES/NO, or 1/0.")

    def strip_cols(df):
        df.columns = df.columns.str.strip()
        return df

    emp_df = strip_cols(pd.read_excel(xl, sheet_name="Employees"))
    employees = [
        {
            "id":       str(r["employee"]).strip(),
            "capacity": int(r["capacity"]),
            "skills":   [s.strip() for s in str(r["skills"]).split(",") if s.strip()],
            "rate":     float(r["hourly_rate ($)"]),
            "type":     str(r.get("employee_type", "")).strip(),
        }
        for _, r in emp_df.iterrows()
    ]

    proj_df = strip_cols(pd.read_excel(xl, sheet_name="Projects"))
    projects = [
        {
            "id":          str(r["project"]).strip(),
            "reimbursable": to_bool(r["reimbursable"]),
            "maxTotal":    None if pd.isna(r["max_total"]) else int(r["max_total"]),
            "budget":      None if pd.isna(r["budget ($)"]) else float(r["budget ($)"]),
        }
        for _, r in proj_df.iterrows()
    ]

    task_df = strip_cols(pd.read_excel(xl, sheet_name="Tasks"))
    tasks = [
        {
            "project":  str(r["project"]).strip(),
            "id":       str(r["task_id"]).strip(),
            "type":     str(r["task_type"]).strip(),
            "minHours": int(r["min_hours"]),
        }
        for _, r in task_df.iterrows()
    ]

    return employees, projects, tasks


# Solve the workforce allocation as a linear program using PuLP.
#
# Variables:
#   x[i,j]  hours employee i works on task j  (>= 0, only valid skill pairs)
#   s[j]    unmet hours for task j             (>= 0)
#
# Objective: minimize sum(w_j * s_j)
#   w_j = 2 for reimbursable project tasks, 1 otherwise (covers billable work first)
#
# Constraints:
#   1. sum_j x[i,j] <= capacity[i]                   (employee capacity)
#   2. sum_i x[i,j] + s[j] >= minHours[j]            (task demand)
#   3. sum_{j in p} x[i,j] <= maxTotal[p]            (project hour cap, reimbursable)
#   4. sum_{j in p} x[i,j] * rate[i] <= budget[p]    (project dollar budget)
def run_optimizer(employees, tasks, projects):
    emp_map  = {e["id"]: e for e in employees}
    task_map = {t["id"]: t for t in tasks}
    proj_map = {p["id"]: p for p in projects}

    # Only create variables for employee-task pairs where the skill matches
    pairs = [(e["id"], t["id"]) for e in employees for t in tasks
             if t["type"] in e["skills"]]

    prob = pulp.LpProblem("wf", pulp.LpMinimize)
    x = {(i, j): pulp.LpVariable(f"x_{i}_{j}", lowBound=0) for i, j in pairs}
    s = {t["id"]: pulp.LpVariable(f"s_{t['id']}", lowBound=0) for t in tasks}

    # Objective:
    # 1) Strongly minimize unmet task hours
    # 2) Then prefer lower total labor cost among equally covered plans
    unmet_penalty = 10000
    prob += pulp.lpSum(
        unmet_penalty * (2 if proj_map.get(task_map[j]["project"], {}).get("reimbursable") else 1) * s[j]
        for j in s
    ) + 0.01 * pulp.lpSum(
        x[(i, j)] * emp_map[i]["rate"] for i, j in pairs
    )

    # Constraint 1: employee capacity
    for e in employees:
        ep = [j for i, j in pairs if i == e["id"]]
        if ep:
            prob += pulp.lpSum(x[(e["id"], j)] for j in ep) <= e["capacity"]

    # Constraint 2: task demand
    for t in tasks:
        tp = [x[(i, t["id"])] for i, j in pairs if j == t["id"]]
        if tp:
            prob += pulp.lpSum(tp) + s[t["id"]] >= t["minHours"]
        else:
            prob += s[t["id"]] >= t["minHours"]

    # Constraint 3: project hour cap (reimbursable only)
    for p in projects:
        if p["reimbursable"] and p["maxTotal"] is not None:
            pp = [(i, j) for i, j in pairs if task_map[j]["project"] == p["id"]]
            if pp:
                prob += pulp.lpSum(x[k] for k in pp) <= p["maxTotal"]

    # Constraint 4: project dollar budget
    for p in projects:
        if p["budget"] is not None:
            pp = [(i, j) for i, j in pairs if task_map[j]["project"] == p["id"]]
            if pp:
                prob += pulp.lpSum(x[(i, j)] * emp_map[i]["rate"] for i, j in pp) <= p["budget"]

    # Use a fast LP solve configuration to keep runtime responsive in Streamlit.
    prob.solve(pulp.PULP_CBC_CMD(msg=0, threads=1))

    # Collect results
    load    = {e["id"]: 0.0 for e in employees}
    p_hours = {p["id"]: 0.0 for p in projects}
    p_cost  = {p["id"]: 0.0 for p in projects}
    ta      = {t["id"]: {} for t in tasks}

    for i, j in pairs:
        v = pulp.value(x[(i, j)])
        if v and v > 0.01:
            ta[j][i] = v
            load[i]                          += v
            p_hours[task_map[j]["project"]]  += v
            p_cost[task_map[j]["project"]]   += v * emp_map[i]["rate"]

    asgn = {}
    for t in tasks:
        assigned = ta[t["id"]]
        if not assigned:
            asgn[t["id"]] = {"employee": None, "hours": 0, "cost": 0.0, "all_assigned": {}}
        else:
            primary = max(assigned, key=assigned.get)
            hours   = sum(assigned.values())
            task_cost = sum(h * emp_map[eid]["rate"] for eid, h in assigned.items())
            asgn[t["id"]] = {
                "employee": primary,
                "hours":    round(hours, 1),
                "partial":  hours < t["minHours"] - 0.5,
                "cost":     round(task_cost, 2),
                "all_assigned": {eid: round(h, 2) for eid, h in assigned.items()},
            }

    return {
        "asgn":    asgn,
        "load":    {k: round(v, 1) for k, v in load.items()},
        "p_hours": {k: round(v, 1) for k, v in p_hours.items()},
        "p_cost":  {k: round(v, 2) for k, v in p_cost.items()},
        "status":  pulp.LpStatus[prob.status],
    }


# Page header
st.markdown("## Workforce Optimizer")
st.caption("Upload your data template in the sidebar to run the optimizer.")

# Sidebar: downloads and file upload
with st.sidebar:
    st.header("Data Import")
    st.markdown("Upload an Excel file with three sheets: **Employees**, **Projects**, and **Tasks**.")

    try:
        st.download_button(
            "Download blank template",
            data=build_blank_template(),
            file_name="workforce_optimizer_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception:
        st.caption("Template download unavailable.")

    try:
        with open("Capstone Budget Data.xlsx", "rb") as f:
            st.download_button(
                "Download example dataset",
                data=f.read(),
                file_name="Capstone Budget Data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    except Exception:
        st.caption("Example dataset unavailable.")

    st.divider()
    uploaded_file = st.file_uploader("Choose Excel file (.xlsx)", type=["xlsx"])


# Stop here until a file is uploaded
if uploaded_file is None:
    st.info("Upload your Excel template in the sidebar to get started.")
    st.stop()

# Parse, validate, and load the uploaded file
try:
    xl = pd.ExcelFile(uploaded_file)
except Exception as e:
    st.error(f"Could not read file: {e}")
    st.stop()

valid, err = validate_template(xl)
if not valid:
    st.error(f"Invalid template:\n\n{err}")
    st.stop()

try:
    EMPLOYEES, PROJECTS, TASKS = load_data(xl)
except Exception as e:
    st.error(f"Error loading data: {e}")
    st.stop()

st.sidebar.success(f"{len(EMPLOYEES)} employees | {len(PROJECTS)} projects | {len(TASKS)} tasks")

# Run the LP optimizer and cache results so tab switching doesn't re-solve
file_bytes = uploaded_file.getvalue()
cache_key = hashlib.md5(file_bytes).hexdigest()
if st.session_state.get("_cache_key") != cache_key:
    with st.spinner("Running optimizer..."):
        st.session_state["opt_result"]  = run_optimizer(EMPLOYEES, TASKS, PROJECTS)
        st.session_state["_cache_key"] = cache_key

opt    = st.session_state["opt_result"]
asgn   = opt["asgn"]
load   = opt["load"]
p_hours = opt["p_hours"]
p_cost  = opt["p_cost"]

proj_map = {p["id"]: p for p in PROJECTS}
emp_map  = {e["id"]: e for e in EMPLOYEES}

total_cap  = sum(e["capacity"] for e in EMPLOYEES)
total_load = round(sum(load.values()), 1)
util       = round(total_load / total_cap * 100) if total_cap else 0
n_ok   = sum(1 for a in asgn.values() if a.get("employee") and not a.get("partial"))
n_fail = sum(1 for a in asgn.values() if not a.get("employee"))
n_part = sum(1 for a in asgn.values() if a.get("partial"))

# Build per-employee task list for the Employees tab
tasks_by_emp = {e["id"]: [] for e in EMPLOYEES}
for t in TASKS:
    a = asgn.get(t["id"], {})
    for eid, hrs in a.get("all_assigned", {}).items():
        if eid in tasks_by_emp:
            tasks_by_emp[eid].append({
                **t,
                "hours": hrs,
                "partial": a.get("partial", False),
                "is_primary": eid == a.get("employee"),
            })

# Top-level KPIs
c1, c2, c3 = st.columns(3)
c1.metric("Utilization",   f"{util}%")
c2.metric("Tasks Covered", f"{n_ok} / {len(TASKS)}")
c3.metric("Unfilled",      n_fail + n_part)

if opt["status"] not in ("Optimal", "Not Solved"):
    st.warning(f"Solver status: {opt['status']} - results may be incomplete.")

st.divider()

tabs = st.tabs(["Dashboard", "Employees", "Projects", "Assignments"])


# Dashboard tab
with tabs[0]:
    reimb = sum(1 for p in PROJECTS if p["reimbursable"])
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Employees",   len(EMPLOYEES), f"{total_cap} h/week capacity")
    c2.metric("Projects",    len(PROJECTS),  f"{reimb} reimbursable")
    c3.metric("Tasks",       f"{n_ok} / {len(TASKS)}", f"{n_fail + n_part} need attention" if n_fail + n_part else "All covered")
    c4.metric("Utilization", f"{util}%",     f"{total_load} of {total_cap} h used")

    st.subheader("Employee Utilization")
    emp_df = pd.DataFrame([{
        "Employee":      e["id"],
        "Type":          e["type"],
        "Rate ($/h)":    f"${e['rate']:.2f}",
        "Skills":        ", ".join(e["skills"]),
        "Assigned (h)":  load.get(e["id"], 0),
        "Capacity (h)":  e["capacity"],
        "Util %":        f"{round(load.get(e['id'], 0) / e['capacity'] * 100) if e['capacity'] else 0}%",
    } for e in EMPLOYEES])
    st.dataframe(emp_df, use_container_width=True, hide_index=True)

    st.subheader("Skills: supply vs demand")
    st.caption(
        "**Supply (h)** uses task-relevant skills only, then **rounds down** to whole hours "
        "(no overstatement). **vs demand** uses that same supply vs task hours."
    )
    skill_types = sorted(set(t["type"] for t in TASKS))
    task_skill_types = set(skill_types)
    skill_rows = []
    for sk in skill_types:
        dem = sum(t["minHours"] for t in TASKS if t["type"] == sk)
        sup = roster_skill_supply_hours(EMPLOYEES, sk, task_skill_types)
        skill_rows.append({
            "Skill":       sk,
            "Demand (h)":  dem,
            "Supply (h)":  int(sup),
            "vs demand":   format_supply_vs_demand_pct(sup, dem),
        })
    st.dataframe(pd.DataFrame(skill_rows), use_container_width=True, hide_index=True)

    st.subheader("Project Summary")
    proj_sum_rows = []
    for p in PROJECTS:
        dem = sum(t["minHours"] for t in TASKS if t["project"] == p["id"])
        asn = p_hours.get(p["id"], 0)
        cost = p_cost.get(p["id"], 0)
        over = (p["maxTotal"] and asn > p["maxTotal"]) or (p["budget"] and cost > p["budget"])
        status_text = "Over Budget" if over else "Fulfilled" if asn >= dem else "Partial"
        proj_sum_rows.append({
            "Project":       p["id"],
            "Type":          "Reimbursable" if p["reimbursable"] else "Non-Reimb.",
            "Demand (h)":    dem,
            "Assigned (h)":  round(asn, 1),
            "Hour Cap":      f"{p['maxTotal']} h" if p["maxTotal"] else "-",
            "Wage Cost ($)": f"${cost:,.0f}",
            "Budget ($)":    f"${p['budget']:,.0f}" if p["budget"] else "-",
            "Staffed %":     f"{round(asn / dem * 100) if dem else 100}%",
            "Status":        status_text,
            "Signal":        status_tier(status_text),
        })
    df_proj = pd.DataFrame(proj_sum_rows)
    st.dataframe(df_proj.style.map(status_style, subset=["Signal"]), use_container_width=True, hide_index=True)


# Employees tab
with tabs[1]:
    st.subheader("Employee Roster")
    for e in EMPLOYEES:
        hrs  = load.get(e["id"], 0)
        util_e = round(hrs / e["capacity"] * 100) if e["capacity"] else 0
        label = (f"**{e['id']}** [{e['type']}]  |  Skills: {', '.join(e['skills'])}"
                 f"  |  ${e['rate']:.2f}/h  |  {util_e}% utilized ({hrs}/{e['capacity']} h)")
        with st.expander(label):
            et = tasks_by_emp.get(e["id"], [])
            if not et:
                st.caption("No tasks assigned.")
            else:
                st.dataframe(pd.DataFrame([{
                    "Task":         t["id"],
                    "Project":      t["project"],
                    "Skill":        t["type"],
                    "Min Hrs":      t["minHours"],
                    "Assigned Hrs": round(t.get("hours", 0), 1),
                    "Role":         "Primary" if t.get("is_primary") else "Contributor",
                    "Status":       "Partial" if t.get("partial") else "OK",
                } for t in et]), use_container_width=True, hide_index=True)


# Projects tab
with tabs[2]:
    st.subheader("Project Overview")
    for p in PROJECTS:
        pt   = [t for t in TASKS if t["project"] == p["id"]]
        dem  = sum(t["minHours"] for t in pt)
        asn  = p_hours.get(p["id"], 0)
        cost = p_cost.get(p["id"], 0)
        over = (p["maxTotal"] and asn > p["maxTotal"]) or (p["budget"] and cost > p["budget"])
        status_label = "Over Budget" if over else ("Fulfilled" if asn >= dem else "Partial")
        signal = status_tier(status_label)
        header = (f"**{p['id']}**  |  {'Reimbursable' if p['reimbursable'] else 'Non-Reimbursable'}"
                  f"  |  {signal}  |  {status_label}  |  {round(asn, 1)}/{dem} h  |  ${cost:,.0f}")

        with st.expander(header):
            # Plain text (no markdown bold): commas in dollar amounts break **…** in captions.
            if p["maxTotal"]:
                hp = min(999, round(asn / p["maxTotal"] * 100))
                st.write(f"Hour budget: {round(asn, 1)} / {p['maxTotal']} h ({hp}% of cap)")
            if p["budget"]:
                wp = min(999, round(cost / p["budget"] * 100))
                st.write(f"Wage budget: ${cost:,.0f} / ${p['budget']:,.0f} ({wp}% of budget)")

            # Cost breakdown: aggregate hours and cost per employee for this project
            st.markdown(
                '<p style="font-size:1rem;font-weight:600;margin:1rem 0 0.35rem 0;">'
                "Cost breakdown</p>",
                unsafe_allow_html=True,
            )
            cb = {}
            for t in pt:
                a = asgn.get(t["id"], {})
                for eid, hrs in a.get("all_assigned", {}).items():
                    cb.setdefault(eid, {"tasks": 0, "hours": 0.0})
                    cb[eid]["tasks"]  += 1
                    cb[eid]["hours"]  += hrs

            if cb:
                cb_rows = [{
                    "Employee":       eid,
                    "Tasks":          v["tasks"],
                    "Hours":          round(v["hours"], 1),
                    "Rate ($/h)":     f"${emp_map[eid]['rate']:.2f}",
                    "Cost ($)":       f"${v['hours'] * emp_map[eid]['rate']:,.2f}",
                } for eid, v in cb.items()]
                cb_rows.append({
                    "Employee": "TOTAL",
                    "Tasks":    sum(r["Tasks"] for r in cb_rows),
                    "Hours":    round(sum(r["Hours"] for r in cb_rows), 1),
                    "Rate ($/h)": "-",
                    "Cost ($)": f"${cost:,.2f}",
                })
                st.dataframe(pd.DataFrame(cb_rows), use_container_width=True, hide_index=True)
            else:
                st.caption("No tasks assigned to this project.")

            st.markdown(
                '<p style="font-size:1rem;font-weight:600;margin:1rem 0 0.35rem 0;">'
                "Task details</p>",
                unsafe_allow_html=True,
            )
            st.dataframe(pd.DataFrame([{
                "Task":         t["id"],
                "Skill":        t["type"],
                "Min Hrs":      t["minHours"],
                "Assigned Hrs": round(asgn.get(t["id"], {}).get("hours", 0), 1),
                "Assigned To":  asgn.get(t["id"], {}).get("employee") or "-",
                "Status":       "Partial"    if asgn.get(t["id"], {}).get("partial")
                                else "OK"    if asgn.get(t["id"], {}).get("employee")
                                else "Unassigned",
            } for t in pt]), use_container_width=True, hide_index=True)


# Assignments tab
with tabs[3]:
    st.subheader("Full Task Assignment Matrix")
    rows = []
    for t in TASKS:
        a   = asgn.get(t["id"], {})
        eid = a.get("employee")
        hrs = a.get("hours", 0)
        owners = list((a.get("all_assigned") or {}).keys())
        owner_label = ", ".join(owners) if owners else "-"
        if len(owners) > 1 and eid:
            owner_label = f"{eid} (primary) + {len(owners)-1} more"
        cost_val = f"${a.get('cost', 0):,.2f}" if owners else "-"
        rows.append({
            "Task":         t["id"],
            "Project":      f"{t['project']} ({'R' if proj_map.get(t['project'], {}).get('reimbursable') else 'N'})",
            "Skill":        t["type"],
            "Min Hours":    t["minHours"],
            "Assigned Hrs": round(hrs, 1),
            "Assigned To":  owner_label,
            "Cost ($)":     cost_val,
            "Status":       "Failed"    if not eid
                            else "Partial" if a.get("partial")
                            else "OK",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
